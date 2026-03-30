"""表格读写：首行表头，支持 xlsx / xls / csv。"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def norm_header(name: str) -> str:
    t = str(name).strip()
    # 双语表头常见「中文\\n英文」，列名匹配时只用第一行（中文）部分
    if "\n" in t or "\r" in t:
        t = t.replace("\r\n", "\n").replace("\r", "\n").split("\n", 1)[0].strip()
    t = t.lower().replace(" ", "").replace("_", "").replace("\u3000", "")
    return t


def resolve_column(headers: list[str], aliases: list[str]) -> str | None:
    """按别名（中英文）在表头中找实际列名。"""
    index: dict[str, str] = {}
    for h in headers:
        index[norm_header(h)] = h
    for alias in aliases:
        key = norm_header(alias)
        if key in index:
            return index[key]
    return None


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return [], []
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        headers = [h for h in headers if h] or headers
        if not any(headers):
            return [], []
        out: list[dict[str, Any]] = []
        for i, row in enumerate(rows, start=2):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            d: dict[str, Any] = {}
            for j, h in enumerate(headers):
                if not h:
                    continue
                val = row[j] if j < len(row) else None
                d[h] = val
            out.append(d)
        return headers, out
    finally:
        wb.close()


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    lines = text.splitlines()
    if not lines:
        return [], []
    sample = lines[0][:2048]
    delimiter = ","
    if "\t" in sample and sample.count("\t") > sample.count(","):
        delimiter = "\t"
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [str(c).strip() for c in rows[0]]
    out: list[dict[str, Any]] = []
    for i, parts in enumerate(rows[1:], start=2):
        if not parts or all(not str(p).strip() for p in parts):
            continue
        d: dict[str, Any] = {}
        for j, h in enumerate(headers):
            if not h:
                continue
            d[h] = parts[j] if j < len(parts) else ""
        out.append(d)
    return headers, out


def load_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _read_csv(path)
    if suf in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if suf == ".xls":
        raise ValueError("不支持 .xls，请另存为 .xlsx 或 .csv 后上传")
    raise ValueError(f"不支持的文件类型: {suf}")


def write_result_xlsx(
    path: Path,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(path)
    logger.info("已写入结果表: %s 行=%s", path, len(rows))
